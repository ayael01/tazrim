from datetime import date
from io import BytesIO
from collections import defaultdict
from decimal import Decimal
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, func, or_
from sqlalchemy.orm import Session, aliased

from app.db.models import (
    BankActivity,
    BankActivityCategory,
    BankActivityImportBatch,
    BankPayee,
    BankPayeeCategoryMap,
)
from app.db.session import get_db
from app.schemas.bank import (
    BankActivitiesExportRequest,
    BankActivityList,
    BankActivityOut,
    BankActivityUpdate,
)

router = APIRouter()


def _build_bank_activities_query(
    db: Session,
    *,
    year: Optional[int] = None,
    month: Optional[int] = None,
    category_id: Optional[int] = None,
    category_name: Optional[str] = None,
    payee_id: Optional[int] = None,
    import_batch_id: Optional[int] = None,
    direction: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    q: Optional[str] = None,
    include_source_filename: bool = False,
):
    manual_category = aliased(BankActivityCategory)
    category_label = func.coalesce(
        manual_category.name,
        BankActivity.raw_category_text,
        BankActivityCategory.name,
        "Uncategorized",
    )
    category_id_value = func.coalesce(
        manual_category.id, BankPayeeCategoryMap.category_id
    )
    query_columns = [
        BankActivity,
        BankPayee.display_name.label("payee_name"),
        category_id_value.label("category_id"),
        category_label.label("category_name"),
    ]
    if include_source_filename:
        query_columns.append(BankActivityImportBatch.source_filename.label("source_filename"))

    query = (
        db.query(*query_columns)
        .outerjoin(BankPayee, BankActivity.payee_id == BankPayee.id)
        .outerjoin(BankPayeeCategoryMap, BankPayee.id == BankPayeeCategoryMap.payee_id)
        .outerjoin(
            BankActivityCategory,
            BankPayeeCategoryMap.category_id == BankActivityCategory.id,
        )
        .outerjoin(manual_category, BankActivity.manual_category_id == manual_category.id)
        .order_by(BankActivity.activity_date.desc(), BankActivity.id.desc())
    )
    if include_source_filename:
        query = query.outerjoin(
            BankActivityImportBatch,
            BankActivity.import_batch_id == BankActivityImportBatch.id,
        )

    if year:
        query = query.filter(func.extract("year", BankActivity.activity_date) == year)
    if month:
        query = query.filter(func.extract("month", BankActivity.activity_date) == month)
    if category_id:
        query = query.filter(
            or_(
                BankActivity.manual_category_id == category_id,
                and_(
                    BankActivity.manual_category_id.is_(None),
                    BankPayeeCategoryMap.category_id == category_id,
                ),
            )
        )
    if category_name:
        query = query.filter(category_label == category_name)
    if payee_id:
        query = query.filter(BankActivity.payee_id == payee_id)
    if import_batch_id:
        query = query.filter(BankActivity.import_batch_id == import_batch_id)
    if date_from:
        query = query.filter(BankActivity.activity_date >= date_from)
    if date_to:
        query = query.filter(BankActivity.activity_date <= date_to)
    if q:
        query = query.filter(
            or_(
                BankActivity.description.ilike(f"%{q}%"),
                BankActivity.payee_raw.ilike(f"%{q}%"),
                BankPayee.display_name.ilike(f"%{q}%"),
                BankActivity.reference.ilike(f"%{q}%"),
            )
        )
    if direction == "income":
        query = query.filter(BankActivity.credit.isnot(None))
    elif direction == "expense":
        query = query.filter(BankActivity.debit.isnot(None))

    return query


@router.get("", response_model=BankActivityList)
def list_bank_activities(
    limit: int = Query(200, ge=1, le=1000),
    offset: int = Query(0, ge=0),
    year: Optional[int] = Query(None, ge=2000, le=2100),
    month: Optional[int] = Query(None, ge=1, le=12),
    category_id: Optional[int] = Query(None, ge=1),
    category_name: Optional[str] = Query(None, min_length=1),
    payee_id: Optional[int] = Query(None, ge=1),
    import_batch_id: Optional[int] = Query(None, ge=1),
    direction: Optional[str] = Query(None, pattern="^(income|expense)$"),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    q: Optional[str] = Query(None, min_length=1),
    db: Session = Depends(get_db),
) -> BankActivityList:
    query = _build_bank_activities_query(
        db,
        year=year,
        month=month,
        category_id=category_id,
        category_name=category_name,
        payee_id=payee_id,
        import_batch_id=import_batch_id,
        direction=direction,
        date_from=date_from,
        date_to=date_to,
        q=q,
    )

    total = (
        query.order_by(None)
        .with_entities(func.count(func.distinct(BankActivity.id)))
        .scalar()
        or 0
    )

    rows = query.limit(limit).offset(offset).all()
    items = []
    for activity, payee_name, category_id_value, category_name in rows:
        items.append(
            BankActivityOut(
                id=activity.id,
                activity_date=activity.activity_date.isoformat(),
                value_date=activity.value_date.isoformat() if activity.value_date else None,
                description=activity.description,
                reference=activity.reference,
                debit=activity.debit,
                credit=activity.credit,
                balance=activity.balance,
                currency=activity.currency,
                payee_name=payee_name,
                category_id=category_id_value,
                category_name=category_name,
            )
        )

    return BankActivityList(total=total, items=items)


@router.post("/export")
def export_bank_activities(
    payload: BankActivitiesExportRequest,
    db: Session = Depends(get_db),
):
    try:
        activity_date_from = (
            date.fromisoformat(payload.date_from)
            if payload.scope == "filtered" and payload.date_from
            else None
        )
        activity_date_to = (
            date.fromisoformat(payload.date_to)
            if payload.scope == "filtered" and payload.date_to
            else None
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid date format for export") from exc
    query = _build_bank_activities_query(
        db,
        category_id=payload.category_id if payload.scope == "filtered" else None,
        direction=payload.direction if payload.scope == "filtered" else None,
        date_from=activity_date_from,
        date_to=activity_date_to,
        q=payload.q if payload.scope == "filtered" else None,
        include_source_filename=True,
    )
    rows = query.all()

    records = []
    for activity, payee_name, _category_id, category_name, source_filename in rows:
        records.append(
            {
                "activity_date": activity.activity_date,
                "value_date": activity.value_date,
                "description": activity.description,
                "reference": activity.reference or "",
                "payee": payee_name or activity.payee_raw,
                "category": category_name or "Uncategorized",
                "debit": activity.debit,
                "credit": activity.credit,
                "balance": activity.balance,
                "currency": activity.currency or "ILS",
                "direction": "Income" if activity.credit is not None else "Expense",
                "raw_category_text": activity.raw_category_text or "",
                "manual_override": "Yes" if activity.manual_category_id is not None else "No",
                "source_filename": source_filename or "",
            }
        )

    column_defs = {
        "activity_date": ("Activity date", 14),
        "value_date": ("Value date", 14),
        "description": ("Description", 36),
        "reference": ("Reference", 16),
        "payee": ("Payee", 28),
        "category": ("Category", 22),
        "debit": ("Debit", 14),
        "credit": ("Credit", 14),
        "balance": ("Balance", 14),
        "currency": ("Currency", 10),
        "direction": ("Direction", 12),
        "raw_category_text": ("Raw category", 22),
        "manual_override": ("Manual override", 16),
        "source_filename": ("Source file", 24),
    }
    selected_columns = [
        key for key in payload.columns if key in column_defs
    ] or list(column_defs.keys())

    workbook = Workbook()
    activities_sheet = workbook.active
    activities_sheet.title = "Activities"
    activities_sheet.freeze_panes = "A2"
    activities_sheet.auto_filter.ref = f"A1:{get_column_letter(len(selected_columns))}1"

    for column_index, key in enumerate(selected_columns, start=1):
        header, width = column_defs[key]
        cell = activities_sheet.cell(row=1, column=column_index, value=header)
        cell.font = Font(bold=True)
        activities_sheet.column_dimensions[get_column_letter(column_index)].width = width

    for row_index, record in enumerate(records, start=2):
        for column_index, key in enumerate(selected_columns, start=1):
            value = record[key]
            cell = activities_sheet.cell(row=row_index, column=column_index, value=value)
            if key in {"activity_date", "value_date"} and value is not None:
                cell.number_format = "yyyy-mm-dd"
            if key in {"debit", "credit", "balance"} and value is not None:
                cell.number_format = "#,##0.00"

    income_total = sum((record["credit"] or 0) for record in records)
    expense_total = sum((record["debit"] or 0) for record in records)
    net_total = income_total - expense_total

    if payload.include_summary:
        summary_sheet = workbook.create_sheet("Summary")
        summary_rows = [
            ("Rows", len(records)),
            ("Income total", income_total),
            ("Expense total", expense_total),
            ("Net", net_total),
            ("Scope", payload.scope),
            ("From", payload.date_from or ""),
            ("To", payload.date_to or ""),
            ("Search", payload.q or ""),
            ("Direction", payload.direction or "all"),
        ]
        for i, (label, value) in enumerate(summary_rows, start=1):
            label_cell = summary_sheet.cell(row=i, column=1, value=label)
            label_cell.font = Font(bold=True)
            value_cell = summary_sheet.cell(row=i, column=2, value=value)
            if isinstance(value, (int, float, Decimal)):
                value_cell.number_format = "#,##0.00"
        summary_sheet.column_dimensions["A"].width = 20
        summary_sheet.column_dimensions["B"].width = 26

    if payload.include_by_category:
        grouped = defaultdict(lambda: {"count": 0, "debit": 0, "credit": 0})
        for record in records:
            bucket = grouped[record["category"]]
            bucket["count"] += 1
            bucket["debit"] += record["debit"] or 0
            bucket["credit"] += record["credit"] or 0
        category_sheet = workbook.create_sheet("By Category")
        headers = ["Category", "Count", "Debit", "Credit", "Net"]
        for column, header in enumerate(headers, start=1):
            cell = category_sheet.cell(row=1, column=column, value=header)
            cell.font = Font(bold=True)
        for row_number, (name, totals) in enumerate(
            sorted(grouped.items(), key=lambda item: (item[1]["debit"] + item[1]["credit"]), reverse=True),
            start=2,
        ):
            category_sheet.cell(row=row_number, column=1, value=name)
            category_sheet.cell(row=row_number, column=2, value=totals["count"])
            category_sheet.cell(row=row_number, column=3, value=totals["debit"]).number_format = "#,##0.00"
            category_sheet.cell(row=row_number, column=4, value=totals["credit"]).number_format = "#,##0.00"
            category_sheet.cell(
                row=row_number, column=5, value=totals["credit"] - totals["debit"]
            ).number_format = "#,##0.00"
        category_sheet.column_dimensions["A"].width = 24
        category_sheet.column_dimensions["B"].width = 10
        category_sheet.column_dimensions["C"].width = 14
        category_sheet.column_dimensions["D"].width = 14
        category_sheet.column_dimensions["E"].width = 14

    if payload.include_by_payee:
        grouped = defaultdict(lambda: {"count": 0, "debit": 0, "credit": 0})
        for record in records:
            bucket = grouped[record["payee"]]
            bucket["count"] += 1
            bucket["debit"] += record["debit"] or 0
            bucket["credit"] += record["credit"] or 0
        payee_sheet = workbook.create_sheet("By Payee")
        headers = ["Payee", "Count", "Debit", "Credit", "Net"]
        for column, header in enumerate(headers, start=1):
            cell = payee_sheet.cell(row=1, column=column, value=header)
            cell.font = Font(bold=True)
        for row_number, (name, totals) in enumerate(
            sorted(grouped.items(), key=lambda item: (item[1]["debit"] + item[1]["credit"]), reverse=True),
            start=2,
        ):
            payee_sheet.cell(row=row_number, column=1, value=name)
            payee_sheet.cell(row=row_number, column=2, value=totals["count"])
            payee_sheet.cell(row=row_number, column=3, value=totals["debit"]).number_format = "#,##0.00"
            payee_sheet.cell(row=row_number, column=4, value=totals["credit"]).number_format = "#,##0.00"
            payee_sheet.cell(
                row=row_number, column=5, value=totals["credit"] - totals["debit"]
            ).number_format = "#,##0.00"
        payee_sheet.column_dimensions["A"].width = 30
        payee_sheet.column_dimensions["B"].width = 10
        payee_sheet.column_dimensions["C"].width = 14
        payee_sheet.column_dimensions["D"].width = 14
        payee_sheet.column_dimensions["E"].width = 14

    if payload.include_monthly_trend:
        grouped = defaultdict(lambda: {"debit": 0, "credit": 0})
        for record in records:
            month_key = record["activity_date"].strftime("%Y-%m")
            bucket = grouped[month_key]
            bucket["debit"] += record["debit"] or 0
            bucket["credit"] += record["credit"] or 0
        trend_sheet = workbook.create_sheet("Monthly Trend")
        headers = ["Month", "Debit", "Credit", "Net"]
        for column, header in enumerate(headers, start=1):
            cell = trend_sheet.cell(row=1, column=column, value=header)
            cell.font = Font(bold=True)
        for row_number, month_key in enumerate(sorted(grouped.keys()), start=2):
            totals = grouped[month_key]
            trend_sheet.cell(row=row_number, column=1, value=month_key)
            trend_sheet.cell(row=row_number, column=2, value=totals["debit"]).number_format = "#,##0.00"
            trend_sheet.cell(row=row_number, column=3, value=totals["credit"]).number_format = "#,##0.00"
            trend_sheet.cell(
                row=row_number, column=4, value=totals["credit"] - totals["debit"]
            ).number_format = "#,##0.00"
        trend_sheet.column_dimensions["A"].width = 12
        trend_sheet.column_dimensions["B"].width = 14
        trend_sheet.column_dimensions["C"].width = 14
        trend_sheet.column_dimensions["D"].width = 14

    filename = (payload.filename or "bank_activities_export").strip() or "bank_activities_export"
    if not filename.lower().endswith(".xlsx"):
        filename = f"{filename}.xlsx"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.patch("/{activity_id}", response_model=BankActivityOut)
def update_bank_activity(
    activity_id: int,
    payload: BankActivityUpdate,
    db: Session = Depends(get_db),
) -> BankActivityOut:
    activity = db.query(BankActivity).filter(BankActivity.id == activity_id).first()
    if not activity:
        raise HTTPException(status_code=404, detail="Activity not found")

    if payload.category_id is None:
        activity.manual_category_id = None
    else:
        category = (
            db.query(BankActivityCategory)
            .filter(BankActivityCategory.id == payload.category_id)
            .first()
        )
        if not category:
            raise HTTPException(status_code=404, detail="Category not found")
        activity.manual_category_id = category.id

    db.commit()

    manual_category = aliased(BankActivityCategory)
    category_label = func.coalesce(
        manual_category.name,
        activity.raw_category_text,
        BankActivityCategory.name,
        "Uncategorized",
    )
    category_id_value = func.coalesce(
        manual_category.id, BankPayeeCategoryMap.category_id
    )
    row = (
        db.query(
            BankActivity,
            BankPayee.display_name.label("payee_name"),
            category_id_value.label("category_id"),
            category_label.label("category_name"),
        )
        .outerjoin(BankPayee, BankActivity.payee_id == BankPayee.id)
        .outerjoin(BankPayeeCategoryMap, BankPayee.id == BankPayeeCategoryMap.payee_id)
        .outerjoin(
            BankActivityCategory,
            BankPayeeCategoryMap.category_id == BankActivityCategory.id,
        )
        .outerjoin(manual_category, BankActivity.manual_category_id == manual_category.id)
        .filter(BankActivity.id == activity_id)
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Activity not found")

    activity, payee_name, category_id_value, category_name = row
    return BankActivityOut(
        id=activity.id,
        activity_date=activity.activity_date.isoformat(),
        value_date=activity.value_date.isoformat() if activity.value_date else None,
        description=activity.description,
        reference=activity.reference,
        debit=activity.debit,
        credit=activity.credit,
        balance=activity.balance,
        currency=activity.currency,
        payee_name=payee_name,
        category_id=category_id_value,
        category_name=category_name,
    )

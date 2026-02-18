from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, func, or_
from sqlalchemy.orm import Session, aliased

from app.db.models import (
    CardAccount,
    Category,
    ImportBatch,
    Merchant,
    MerchantCategoryMap,
    Transaction,
)
from app.db.session import get_db
from app.schemas.transactions import (
    TransactionsExportRequest,
    TransactionList,
    TransactionMonthList,
    TransactionOut,
    TransactionUpdate,
)

router = APIRouter()


def _build_transactions_query(
    db: Session,
    *,
    category_id: Optional[int] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    q: Optional[str] = None,
    merchant_id: Optional[int] = None,
    include_source_filename: bool = False,
):
    manual_category = aliased(Category)
    category_label = func.coalesce(
        manual_category.name,
        Category.name,
        "Uncategorized",
    )
    category_id_value = func.coalesce(
        manual_category.id, MerchantCategoryMap.category_id
    )
    query_columns = [
        Transaction,
        Merchant.display_name.label("merchant_display"),
        CardAccount.name.label("card_account_name"),
        category_id_value.label("category_id"),
        category_label.label("category_name"),
    ]
    if include_source_filename:
        query_columns.append(ImportBatch.source_filename.label("source_filename"))

    query = (
        db.query(*query_columns)
        .outerjoin(Merchant, Transaction.merchant_id == Merchant.id)
        .outerjoin(MerchantCategoryMap, Merchant.id == MerchantCategoryMap.merchant_id)
        .outerjoin(Category, MerchantCategoryMap.category_id == Category.id)
        .outerjoin(manual_category, Transaction.manual_category_id == manual_category.id)
        .outerjoin(CardAccount, Transaction.card_account_id == CardAccount.id)
        .order_by(Transaction.transaction_date.desc(), Transaction.id.desc())
    )
    if include_source_filename:
        query = query.outerjoin(ImportBatch, Transaction.import_batch_id == ImportBatch.id)

    if category_id:
        query = query.filter(
            or_(
                Transaction.manual_category_id == category_id,
                and_(
                    Transaction.manual_category_id.is_(None),
                    MerchantCategoryMap.category_id == category_id,
                ),
            )
        )
    if date_from:
        query = query.filter(Transaction.transaction_date >= date_from)
    if date_to:
        query = query.filter(Transaction.transaction_date <= date_to)
    if q:
        query = query.filter(
            or_(
                Transaction.merchant_raw.ilike(f"%{q}%"),
                Merchant.display_name.ilike(f"%{q}%"),
            )
        )
    if merchant_id:
        query = query.filter(Transaction.merchant_id == merchant_id)

    return query


def _effective_amount_and_currency(record: dict, currency_mode: str) -> tuple[Decimal, str]:
    transaction_amount = record["transaction_amount"] or Decimal("0")
    charged_amount = record["charged_amount"]
    transaction_currency = record["transaction_currency"] or "ILS"
    charged_currency = record["charged_currency"] or transaction_currency

    if currency_mode == "original":
        return transaction_amount, transaction_currency
    if currency_mode == "charged":
        if charged_amount is not None:
            return charged_amount, charged_currency
        return transaction_amount, transaction_currency
    if charged_amount is not None:
        return charged_amount, charged_currency
    return transaction_amount, transaction_currency


@router.get("", response_model=TransactionList)
def list_transactions(
    limit: int = Query(200, ge=1, le=1000),
    offset: int = Query(0, ge=0),
    import_batch_id: Optional[int] = Query(None, ge=1),
    category_id: Optional[int] = Query(None, ge=1),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    q: Optional[str] = Query(None, min_length=1),
    db: Session = Depends(get_db),
) -> TransactionList:
    query = _build_transactions_query(
        db,
        category_id=category_id,
        date_from=date_from,
        date_to=date_to,
        q=q,
    )
    if import_batch_id:
        query = query.filter(Transaction.import_batch_id == import_batch_id)

    total = (
        query.order_by(None)
        .with_entities(func.count(func.distinct(Transaction.id)))
        .scalar()
        or 0
    )

    query = query.limit(limit).offset(offset)

    items = []
    for transaction, _merchant_display, _card_name, category_id_value, category_name in query.all():
        items.append(
            TransactionOut(
                id=transaction.id,
                transaction_date=transaction.transaction_date,
                posting_date=transaction.posting_date,
                merchant_raw=transaction.merchant_raw,
                amount=transaction.transaction_amount,
                currency=transaction.transaction_currency,
                charged_amount=transaction.charged_amount,
                charged_currency=transaction.charged_currency,
                manual_category_id=transaction.manual_category_id,
                category_id=category_id_value,
                category_name=category_name,
            )
        )

    return TransactionList(total=total, items=items)


@router.get("/merchant-month", response_model=TransactionMonthList)
def merchant_month_transactions(
    merchant_id: int = Query(..., ge=1),
    year: int = Query(..., ge=2000, le=2100),
    month: int = Query(..., ge=1, le=12),
    db: Session = Depends(get_db),
) -> TransactionMonthList:
    month_start = datetime(year, month, 1)
    month_end = datetime(year + (month == 12), 1 if month == 12 else month + 1, 1)

    query = _build_transactions_query(
        db,
        merchant_id=merchant_id,
        date_from=month_start.date(),
        date_to=(month_end.date()),
    )
    query = query.filter(Transaction.transaction_date < month_end.date())

    items = []
    for transaction, _merchant_display, _card_name, category_id_value, category_name in query.all():
        items.append(
            TransactionOut(
                id=transaction.id,
                transaction_date=transaction.transaction_date,
                posting_date=transaction.posting_date,
                merchant_raw=transaction.merchant_raw,
                amount=transaction.transaction_amount,
                currency=transaction.transaction_currency,
                charged_amount=transaction.charged_amount,
                charged_currency=transaction.charged_currency,
                manual_category_id=transaction.manual_category_id,
                category_id=category_id_value,
                category_name=category_name,
            )
        )

    month_label = f"{year:04d}-{month:02d}"
    return TransactionMonthList(month=month_label, items=items)


@router.post("/export")
def export_transactions(
    payload: TransactionsExportRequest,
    db: Session = Depends(get_db),
):
    try:
        transaction_date_from = (
            date.fromisoformat(payload.date_from)
            if payload.scope == "filtered" and payload.date_from
            else None
        )
        transaction_date_to = (
            date.fromisoformat(payload.date_to)
            if payload.scope == "filtered" and payload.date_to
            else None
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid date format for export") from exc

    query = _build_transactions_query(
        db,
        category_id=payload.category_id if payload.scope == "filtered" else None,
        date_from=transaction_date_from,
        date_to=transaction_date_to,
        q=payload.q if payload.scope == "filtered" else None,
        include_source_filename=True,
    )
    rows = query.all()

    records = []
    for transaction, merchant_display, card_account_name, _category_id, category_name, source_filename in rows:
        records.append(
            {
                "transaction_date": transaction.transaction_date,
                "posting_date": transaction.posting_date,
                "billed_month": (
                    transaction.posting_date.strftime("%Y-%m")
                    if transaction.posting_date
                    else ""
                ),
                "merchant_raw": transaction.merchant_raw,
                "merchant_display": merchant_display or transaction.merchant_raw,
                "category": category_name or "Uncategorized",
                "manual_override": "Yes" if transaction.manual_category_id is not None else "No",
                "transaction_amount": transaction.transaction_amount,
                "transaction_currency": transaction.transaction_currency or "ILS",
                "charged_amount": transaction.charged_amount,
                "charged_currency": (
                    transaction.charged_currency
                    or transaction.transaction_currency
                    or "ILS"
                ),
                "card_account": card_account_name or "",
                "source_filename": source_filename or "",
                "transaction_id": transaction.id,
            }
        )

    for record in records:
        amount, currency = _effective_amount_and_currency(record, payload.currency_mode)
        record["effective_amount"] = amount
        record["effective_currency"] = currency

    column_defs = {
        "transaction_date": ("Date", 14),
        "posting_date": ("Billed date", 14),
        "billed_month": ("Billed month", 12),
        "merchant_raw": ("Merchant (raw)", 30),
        "merchant_display": ("Merchant", 30),
        "category": ("Category", 22),
        "manual_override": ("Manual override", 16),
        "transaction_amount": ("Amount (original)", 16),
        "transaction_currency": ("Original currency", 14),
        "charged_amount": ("Amount (charged)", 16),
        "charged_currency": ("Charged currency", 14),
        "effective_amount": ("Amount (selected mode)", 20),
        "effective_currency": ("Selected currency", 16),
        "card_account": ("Card account", 18),
        "source_filename": ("Source file", 24),
        "transaction_id": ("Transaction id", 16),
    }
    selected_columns = [key for key in payload.columns if key in column_defs] or list(column_defs.keys())

    workbook = Workbook()
    transactions_sheet = workbook.active
    transactions_sheet.title = "Transactions"
    transactions_sheet.freeze_panes = "A2"
    transactions_sheet.auto_filter.ref = f"A1:{get_column_letter(len(selected_columns))}1"

    for column_index, key in enumerate(selected_columns, start=1):
        header, width = column_defs[key]
        cell = transactions_sheet.cell(row=1, column=column_index, value=header)
        cell.font = Font(bold=True)
        transactions_sheet.column_dimensions[get_column_letter(column_index)].width = width

    for row_index, record in enumerate(records, start=2):
        for column_index, key in enumerate(selected_columns, start=1):
            value = record[key]
            cell = transactions_sheet.cell(row=row_index, column=column_index, value=value)
            if key in {"transaction_date", "posting_date"} and value is not None:
                cell.number_format = "yyyy-mm-dd"
            if key in {"transaction_amount", "charged_amount", "effective_amount"} and value is not None:
                cell.number_format = "#,##0.00"

    monthly_totals = defaultdict(lambda: {"count": 0, "amount": Decimal("0")})
    category_totals = defaultdict(lambda: {"count": 0, "amount": Decimal("0")})
    merchant_totals = defaultdict(lambda: {"count": 0, "amount": Decimal("0")})
    billing_cycle_totals = defaultdict(lambda: {"count": 0, "amount": Decimal("0")})
    month_category_totals = defaultdict(lambda: defaultdict(lambda: Decimal("0")))

    for record in records:
        amount = record["effective_amount"] or Decimal("0")
        month_key = record["transaction_date"].strftime("%Y-%m")
        monthly_totals[month_key]["count"] += 1
        monthly_totals[month_key]["amount"] += amount
        category_totals[record["category"]]["count"] += 1
        category_totals[record["category"]]["amount"] += amount
        month_category_totals[month_key][record["category"]] += amount
        merchant_totals[record["merchant_display"]]["count"] += 1
        merchant_totals[record["merchant_display"]]["amount"] += amount
        billing_key = record["billed_month"] or "Unbilled"
        billing_cycle_totals[billing_key]["count"] += 1
        billing_cycle_totals[billing_key]["amount"] += amount

    grand_total = sum((record["effective_amount"] or Decimal("0")) for record in records)
    active_months = len(monthly_totals)
    average_month = (grand_total / active_months) if active_months else Decimal("0")

    highest_month_name = ""
    highest_month_total = Decimal("0")
    lowest_month_name = ""
    lowest_month_total = Decimal("0")
    if monthly_totals:
        highest_month_name, highest_bucket = max(monthly_totals.items(), key=lambda item: item[1]["amount"])
        lowest_month_name, lowest_bucket = min(monthly_totals.items(), key=lambda item: item[1]["amount"])
        highest_month_total = highest_bucket["amount"]
        lowest_month_total = lowest_bucket["amount"]

    if payload.include_summary:
        summary_sheet = workbook.create_sheet("Summary")
        summary_rows = [
            ("Rows", len(records)),
            ("Total spent", grand_total),
            ("Average monthly spent", average_month),
            ("Highest month", highest_month_name),
            ("Highest month spent", highest_month_total),
            ("Lowest month", lowest_month_name),
            ("Lowest month spent", lowest_month_total),
            ("Currency mode", payload.currency_mode),
            ("Scope", payload.scope),
            ("From", payload.date_from or ""),
            ("To", payload.date_to or ""),
            ("Search", payload.q or ""),
        ]
        for i, (label, value) in enumerate(summary_rows, start=1):
            summary_sheet.cell(row=i, column=1, value=label).font = Font(bold=True)
            value_cell = summary_sheet.cell(row=i, column=2, value=value)
            if isinstance(value, (int, float, Decimal)):
                value_cell.number_format = "#,##0.00"
        summary_sheet.column_dimensions["A"].width = 24
        summary_sheet.column_dimensions["B"].width = 30

    if payload.include_monthly_trend:
        trend_sheet = workbook.create_sheet("Monthly Trend")
        headers = ["Month", "Transactions", "Total spent", "Average per transaction"]
        for column, header in enumerate(headers, start=1):
            trend_sheet.cell(row=1, column=column, value=header).font = Font(bold=True)
        for row_number, month_key in enumerate(sorted(monthly_totals.keys()), start=2):
            bucket = monthly_totals[month_key]
            trend_sheet.cell(row=row_number, column=1, value=month_key)
            trend_sheet.cell(row=row_number, column=2, value=bucket["count"])
            trend_sheet.cell(row=row_number, column=3, value=bucket["amount"]).number_format = "#,##0.00"
            avg_value = (bucket["amount"] / bucket["count"]) if bucket["count"] else Decimal("0")
            trend_sheet.cell(row=row_number, column=4, value=avg_value).number_format = "#,##0.00"
        trend_sheet.column_dimensions["A"].width = 12
        trend_sheet.column_dimensions["B"].width = 14
        trend_sheet.column_dimensions["C"].width = 16
        trend_sheet.column_dimensions["D"].width = 24

    if payload.include_monthly_category_matrix:
        matrix_sheet = workbook.create_sheet("Monthly Category Matrix")
        month_keys = sorted(monthly_totals.keys())
        category_names = [
            name
            for name, _bucket in sorted(
                category_totals.items(),
                key=lambda item: item[1]["amount"],
                reverse=True,
            )
        ]

        headers = ["Month", *category_names, "TOTAL"]
        for column, header in enumerate(headers, start=1):
            header_cell = matrix_sheet.cell(row=1, column=column, value=header)
            header_cell.font = Font(bold=True)
            if column == 1:
                matrix_sheet.column_dimensions[get_column_letter(column)].width = 12
            elif column == len(headers):
                matrix_sheet.column_dimensions[get_column_letter(column)].width = 16
            else:
                matrix_sheet.column_dimensions[get_column_letter(column)].width = min(
                    max(len(str(header)) + 6, 14),
                    30,
                )

        matrix_sheet.freeze_panes = "B2"
        matrix_sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        for row_number, month_key in enumerate(month_keys, start=2):
            matrix_sheet.cell(row=row_number, column=1, value=month_key)
            row_total = Decimal("0")
            for category_index, category_name in enumerate(category_names, start=2):
                amount = month_category_totals[month_key].get(category_name, Decimal("0"))
                matrix_sheet.cell(row=row_number, column=category_index, value=amount).number_format = "#,##0.00"
                row_total += amount
            matrix_sheet.cell(row=row_number, column=len(headers), value=row_total).number_format = "#,##0.00"

        total_row = len(month_keys) + 2
        total_label_cell = matrix_sheet.cell(row=total_row, column=1, value="Total")
        total_label_cell.font = Font(bold=True)
        for category_index, category_name in enumerate(category_names, start=2):
            total_cell = matrix_sheet.cell(
                row=total_row,
                column=category_index,
                value=category_totals[category_name]["amount"],
            )
            total_cell.font = Font(bold=True)
            total_cell.number_format = "#,##0.00"
        grand_total_cell = matrix_sheet.cell(row=total_row, column=len(headers), value=grand_total)
        grand_total_cell.font = Font(bold=True)
        grand_total_cell.number_format = "#,##0.00"

    if payload.include_by_category:
        category_sheet = workbook.create_sheet("By Category")
        headers = ["Category", "Transactions", "Total spent", "Average per transaction"]
        for column, header in enumerate(headers, start=1):
            category_sheet.cell(row=1, column=column, value=header).font = Font(bold=True)
        for row_number, (name, bucket) in enumerate(
            sorted(category_totals.items(), key=lambda item: item[1]["amount"], reverse=True),
            start=2,
        ):
            category_sheet.cell(row=row_number, column=1, value=name)
            category_sheet.cell(row=row_number, column=2, value=bucket["count"])
            category_sheet.cell(row=row_number, column=3, value=bucket["amount"]).number_format = "#,##0.00"
            avg_value = (bucket["amount"] / bucket["count"]) if bucket["count"] else Decimal("0")
            category_sheet.cell(row=row_number, column=4, value=avg_value).number_format = "#,##0.00"
        category_sheet.column_dimensions["A"].width = 26
        category_sheet.column_dimensions["B"].width = 14
        category_sheet.column_dimensions["C"].width = 16
        category_sheet.column_dimensions["D"].width = 24

    if payload.include_by_merchant:
        merchant_sheet = workbook.create_sheet("By Merchant")
        headers = ["Merchant", "Transactions", "Total spent", "Average per transaction"]
        for column, header in enumerate(headers, start=1):
            merchant_sheet.cell(row=1, column=column, value=header).font = Font(bold=True)
        for row_number, (name, bucket) in enumerate(
            sorted(merchant_totals.items(), key=lambda item: item[1]["amount"], reverse=True),
            start=2,
        ):
            merchant_sheet.cell(row=row_number, column=1, value=name)
            merchant_sheet.cell(row=row_number, column=2, value=bucket["count"])
            merchant_sheet.cell(row=row_number, column=3, value=bucket["amount"]).number_format = "#,##0.00"
            avg_value = (bucket["amount"] / bucket["count"]) if bucket["count"] else Decimal("0")
            merchant_sheet.cell(row=row_number, column=4, value=avg_value).number_format = "#,##0.00"
        merchant_sheet.column_dimensions["A"].width = 34
        merchant_sheet.column_dimensions["B"].width = 14
        merchant_sheet.column_dimensions["C"].width = 16
        merchant_sheet.column_dimensions["D"].width = 24

    if payload.include_billing_cycle:
        billing_sheet = workbook.create_sheet("Billing Cycles")
        headers = ["Billed month", "Transactions", "Total charged"]
        for column, header in enumerate(headers, start=1):
            billing_sheet.cell(row=1, column=column, value=header).font = Font(bold=True)
        for row_number, month_key in enumerate(sorted(billing_cycle_totals.keys()), start=2):
            bucket = billing_cycle_totals[month_key]
            billing_sheet.cell(row=row_number, column=1, value=month_key)
            billing_sheet.cell(row=row_number, column=2, value=bucket["count"])
            billing_sheet.cell(row=row_number, column=3, value=bucket["amount"]).number_format = "#,##0.00"
        billing_sheet.column_dimensions["A"].width = 16
        billing_sheet.column_dimensions["B"].width = 14
        billing_sheet.column_dimensions["C"].width = 16

    if payload.include_exceptions:
        exceptions_sheet = workbook.create_sheet("Exceptions")
        headers = [
            "Date",
            "Merchant",
            "Category",
            "Issue",
            "Amount (original)",
            "Amount (charged)",
            "Currency pair",
            "Transaction id",
        ]
        for column, header in enumerate(headers, start=1):
            exceptions_sheet.cell(row=1, column=column, value=header).font = Font(bold=True)
        row_number = 2
        for record in records:
            issues = []
            if record["category"] == "Uncategorized":
                issues.append("Uncategorized")
            if record["charged_amount"] is None:
                issues.append("Missing charged amount")
            if (
                record["charged_amount"] is not None
                and record["transaction_amount"] is not None
                and record["charged_amount"] != record["transaction_amount"]
            ):
                issues.append("Amount differs from original")
            if record["charged_currency"] != record["transaction_currency"]:
                issues.append("Currency differs")
            if not issues:
                continue

            exceptions_sheet.cell(row=row_number, column=1, value=record["transaction_date"]).number_format = "yyyy-mm-dd"
            exceptions_sheet.cell(row=row_number, column=2, value=record["merchant_display"])
            exceptions_sheet.cell(row=row_number, column=3, value=record["category"])
            exceptions_sheet.cell(row=row_number, column=4, value=", ".join(issues))
            exceptions_sheet.cell(row=row_number, column=5, value=record["transaction_amount"]).number_format = "#,##0.00"
            exceptions_sheet.cell(row=row_number, column=6, value=record["charged_amount"]).number_format = "#,##0.00"
            exceptions_sheet.cell(
                row=row_number,
                column=7,
                value=f'{record["transaction_currency"]} -> {record["charged_currency"]}',
            )
            exceptions_sheet.cell(row=row_number, column=8, value=record["transaction_id"])
            row_number += 1

        exceptions_sheet.column_dimensions["A"].width = 14
        exceptions_sheet.column_dimensions["B"].width = 28
        exceptions_sheet.column_dimensions["C"].width = 24
        exceptions_sheet.column_dimensions["D"].width = 38
        exceptions_sheet.column_dimensions["E"].width = 16
        exceptions_sheet.column_dimensions["F"].width = 16
        exceptions_sheet.column_dimensions["G"].width = 16
        exceptions_sheet.column_dimensions["H"].width = 14

    filename = (payload.filename or "card_transactions_export").strip() or "card_transactions_export"
    if not filename.lower().endswith(".xlsx"):
        filename = f"{filename}.xlsx"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.patch("/{transaction_id}", response_model=TransactionOut)
def update_transaction(
    transaction_id: int,
    payload: TransactionUpdate,
    db: Session = Depends(get_db),
) -> TransactionOut:
    transaction = (
        db.query(Transaction).filter(Transaction.id == transaction_id).first()
    )
    if not transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")

    if payload.category_id is None:
        transaction.manual_category_id = None
    else:
        category = (
            db.query(Category).filter(Category.id == payload.category_id).first()
        )
        if not category:
            raise HTTPException(status_code=404, detail="Category not found")
        transaction.manual_category_id = category.id

    db.commit()

    manual_category = aliased(Category)
    category_label = func.coalesce(
        manual_category.name,
        Category.name,
        "Uncategorized",
    )
    category_id_value = func.coalesce(
        manual_category.id, MerchantCategoryMap.category_id
    )
    row = (
        db.query(
            Transaction,
            category_id_value.label("category_id"),
            category_label.label("category_name"),
        )
        .outerjoin(Merchant, Transaction.merchant_id == Merchant.id)
        .outerjoin(MerchantCategoryMap, Merchant.id == MerchantCategoryMap.merchant_id)
        .outerjoin(Category, MerchantCategoryMap.category_id == Category.id)
        .outerjoin(manual_category, Transaction.manual_category_id == manual_category.id)
        .filter(Transaction.id == transaction_id)
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Transaction not found")

    transaction, category_id_value, category_name = row
    return TransactionOut(
        id=transaction.id,
        transaction_date=transaction.transaction_date,
        posting_date=transaction.posting_date,
        merchant_raw=transaction.merchant_raw,
        amount=transaction.transaction_amount,
        currency=transaction.transaction_currency,
        charged_amount=transaction.charged_amount,
        charged_currency=transaction.charged_currency,
        manual_category_id=transaction.manual_category_id,
        category_id=category_id_value,
        category_name=category_name,
    )
